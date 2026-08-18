"""openpyxl-backed workbook output adapter.

This adapter preserves Excel as an output/compatibility surface. It never
uses Excel formulas as CenValue calculation authority and never declares
Microsoft Excel qualification PASS.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from pathlib import Path
import os
import tempfile
import zipfile

from openpyxl import load_workbook

from ...ports.workbook_output import (
    WorkbookGenerationArtifact,
    WorkbookGenerationSourceBinding,
    WorkbookOutputWriter,
    WorkbookScalar,
)
from ..excel.fingerprint import (
    FormulaObservation,
    SheetObservation,
    WorkbookFingerprintObservation,
    match_template_profile,
    normalize_formula,
)
from ..excel.profile import ExternalLinkState, SheetState
from .profile import (
    WorkbookFixedSourceBinding,
    WorkbookOutputProfile,
    WorkbookValueKind,
    WorkbookWriteBinding,
)


class WorkbookOutputError(RuntimeError):
    pass


class UnsupportedWorkbookOutputProfileError(WorkbookOutputError):
    pass


class WorkbookSourceHashMismatchError(WorkbookOutputError):
    pass


class WorkbookWriteContractError(WorkbookOutputError):
    pass


class WorkbookStructuralRegressionError(WorkbookOutputError):
    pass


_FIXED_ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)


def _sha256_file(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024, b""), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _split_cell(cell: str) -> tuple[str, str]:
    sheet, coordinate = cell.rsplit("!", 1)
    return sheet, coordinate


def _sheet_state(value: str) -> SheetState:
    mapping = {
        "visible": SheetState.VISIBLE,
        "hidden": SheetState.HIDDEN,
        "veryHidden": SheetState.VERY_HIDDEN,
    }
    try:
        return mapping[value]
    except KeyError as exc:
        raise WorkbookStructuralRegressionError(
            f"unsupported worksheet state {value!r}"
        ) from exc


def _canonical_cell_value(value: object) -> object:
    if isinstance(value, Decimal):
        return ("DECIMAL", str(value))
    if isinstance(value, datetime):
        return ("DATETIME", value.isoformat())
    if isinstance(value, date):
        return ("DATE", value.isoformat())
    return value


def _cell_value_map(workbook) -> dict[str, object]:
    values: dict[str, object] = {}
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values[f"{worksheet.title}!{cell.coordinate}"] = _canonical_cell_value(
                        cell.value
                    )
    return values


def _formula_cells(workbook) -> frozenset[str]:
    output: set[str] = set()
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    output.add(f"{worksheet.title}!{cell.coordinate}")
    return frozenset(output)


def _external_link_state(workbook) -> ExternalLinkState:
    external_formula_cells: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if (
                    isinstance(value, str)
                    and value.startswith("=")
                    and "[" in value
                    and "]" in value
                ):
                    external_formula_cells.append(
                        f"{worksheet.title}!{cell.coordinate}"
                    )

    package_links = tuple(getattr(workbook, "_external_links", ()) or ())
    if not external_formula_cells and not package_links:
        return ExternalLinkState.NONE
    if set(external_formula_cells) == {"Phieu TTTT!E5"} and len(package_links) <= 1:
        return ExternalLinkState.KNOWN_STALE_SELF_REFERENCE
    return ExternalLinkState.UNKNOWN


def _observe(workbook, profile: WorkbookOutputProfile, filename: str):
    template = profile.template_profile
    formulas: list[FormulaObservation] = []
    for signature in template.formula_signatures:
        sheet_name, coordinate = _split_cell(signature.cell)
        if sheet_name not in workbook.sheetnames:
            continue
        value = workbook[sheet_name][coordinate].value
        if isinstance(value, str) and value.startswith("="):
            formulas.append(FormulaObservation(signature.cell, value))

    return WorkbookFingerprintObservation(
        sheets=tuple(
            SheetObservation(worksheet.title, _sheet_state(worksheet.sheet_state))
            for worksheet in workbook.worksheets
        ),
        formulas=tuple(formulas),
        external_link_state=_external_link_state(workbook),
        controls=frozenset(),
        filename=filename,
    )


def _coerce_value(binding, value: WorkbookScalar):
    if value is None:
        if binding.required:
            raise WorkbookWriteContractError(
                f"required workbook value {binding.source_key} is missing"
            )
        return None
    if binding.value_kind is WorkbookValueKind.TEXT:
        return str(value)
    if isinstance(value, bool):
        raise WorkbookWriteContractError(
            f"numeric workbook value {binding.source_key} may not be bool"
        )
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise WorkbookWriteContractError(
            f"{binding.source_key} is not a valid decimal value"
        ) from exc
    if not result.is_finite():
        raise WorkbookWriteContractError(
            f"{binding.source_key} must be finite"
        )
    return result


def _fixed_value_matches(
    binding: WorkbookFixedSourceBinding,
    *,
    expected: WorkbookScalar,
    actual: object,
) -> bool:
    expected_value = _coerce_value(binding, expected)
    if binding.value_kind is WorkbookValueKind.TEXT:
        return actual is not None and str(actual) == expected_value
    if actual is None or isinstance(actual, bool):
        return False
    try:
        actual_decimal = Decimal(str(actual))
        tolerance = Decimal(binding.numeric_tolerance)
    except (InvalidOperation, ValueError):
        return False
    return actual_decimal.is_finite() and abs(actual_decimal - expected_value) <= tolerance


def _verify_fixed_source_bindings(
    cached_workbook,
    profile: WorkbookOutputProfile,
    values,
) -> None:
    for binding in profile.fixed_source_bindings:
        if binding.source_key not in values or values[binding.source_key] is None:
            if binding.required:
                raise WorkbookWriteContractError(
                    f"required fixed source value {binding.source_key} is missing"
                )
            continue
        sheet_name, coordinate = _split_cell(binding.cell)
        if sheet_name not in cached_workbook.sheetnames:
            raise WorkbookStructuralRegressionError(
                f"fixed source worksheet missing: {sheet_name}"
            )
        actual = cached_workbook[sheet_name][coordinate].value
        if not _fixed_value_matches(
            binding,
            expected=values[binding.source_key],
            actual=actual,
        ):
            raise WorkbookWriteContractError(
                f"current canonical value {binding.source_key} is incompatible with "
                f"read-only exemplar cell {binding.cell}"
            )


def _normalize_xlsx_zip(path: Path) -> None:
    normalized = path.with_suffix(path.suffix + ".normalized")
    try:
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
            normalized, "w", allowZip64=True
        ) as target:
            for info in sorted(source.infolist(), key=lambda item: item.filename):
                data = source.read(info.filename)
                clean = zipfile.ZipInfo(info.filename, _FIXED_ZIP_TIMESTAMP)
                clean.compress_type = info.compress_type
                clean.comment = info.comment
                clean.extra = info.extra
                clean.internal_attr = info.internal_attr
                clean.external_attr = info.external_attr
                clean.create_system = info.create_system
                target.writestr(clean, data)
        os.replace(normalized, path)
    finally:
        if normalized.exists():
            normalized.unlink()


def _new_owned_temporary(output: Path) -> Path:
    descriptor, name = tempfile.mkstemp(
        prefix=f".{output.stem}.",
        suffix=".tmp.xlsx",
        dir=output.parent,
    )
    os.close(descriptor)
    return Path(name)


def _publish_new_artifact(temporary: Path, output: Path) -> None:
    """Atomically create output without replacing an existing destination.

    The temporary file lives in the destination directory, so a hard-link
    publication is a single-filesystem create-if-absent operation. Unlike
    os.replace(), os.link() fails when the destination name already exists.
    """

    try:
        os.link(temporary, output)
    except FileExistsError as exc:
        raise WorkbookWriteContractError(
            "output_path became occupied before artifact publication"
        ) from exc
    except OSError as exc:
        raise WorkbookWriteContractError(
            "generated workbook could not be atomically published without overwrite"
        ) from exc


def _cleanup_owned_after_publish_failure(temporary: Path, output: Path) -> None:
    """Clean only names still proven to reference this attempt's temp file."""

    try:
        if temporary.exists() and output.exists() and os.path.samefile(temporary, output):
            output.unlink()
    except OSError:
        # Never fall back to deleting an unproven destination path.
        pass
    try:
        if temporary.exists():
            temporary.unlink()
    except OSError:
        pass


class OpenPyxlWorkbookOutputWriter(WorkbookOutputWriter):
    def __init__(self, profiles: tuple[WorkbookOutputProfile, ...]) -> None:
        self._profiles = profiles

    def _profile(self, profile_id: str, profile_version: str) -> WorkbookOutputProfile:
        matches = [
            item
            for item in self._profiles
            if item.profile_id == profile_id and item.profile_version == profile_version
        ]
        if len(matches) != 1:
            raise UnsupportedWorkbookOutputProfileError(
                f"unsupported workbook output profile: {profile_id}@{profile_version}"
            )
        return matches[0]

    @staticmethod
    def _verify_consumers(workbook, profile: WorkbookOutputProfile) -> None:
        for consumer in profile.output_consumers:
            sheet_name, coordinate = _split_cell(consumer.cell)
            if sheet_name not in workbook.sheetnames:
                raise WorkbookStructuralRegressionError(
                    f"output consumer sheet missing: {sheet_name}"
                )
            actual = workbook[sheet_name][coordinate].value
            if not isinstance(actual, str) or not actual.startswith("="):
                raise WorkbookStructuralRegressionError(
                    f"output consumer {consumer.cell} is no longer a formula"
                )
            if normalize_formula(actual) != normalize_formula(consumer.expected_formula):
                raise WorkbookStructuralRegressionError(
                    f"output consumer formula mismatch at {consumer.cell}"
                )

    def generate(
        self,
        *,
        profile_id: str,
        profile_version: str,
        template_path: str,
        output_path: str,
        values,
        source_binding: WorkbookGenerationSourceBinding,
        generated_at: str,
    ) -> WorkbookGenerationArtifact:
        profile = self._profile(profile_id, profile_version)
        source = Path(template_path).expanduser().resolve()
        output = Path(output_path).expanduser().resolve()
        if source == output:
            raise WorkbookWriteContractError(
                "reference workbook may not be edited in place"
            )
        if not source.is_file():
            raise WorkbookWriteContractError("template_path must be an existing file")
        if output.exists():
            raise WorkbookWriteContractError("output_path must not already exist")
        if source.suffix.lower() != ".xlsx" or output.suffix.lower() != ".xlsx":
            raise WorkbookWriteContractError("E1-PR-005 supports .xlsx artifacts only")

        source_sha = _sha256_file(source)
        if source_sha != profile.source_exemplar_sha256:
            raise WorkbookSourceHashMismatchError(
                "source workbook SHA-256 does not match the supported exemplar"
            )

        required_bindings = (
            *profile.write_bindings,
            *profile.fixed_source_bindings,
            *profile.compatibility_bindings,
        )
        missing = sorted(
            item.source_key
            for item in required_bindings
            if item.required and (item.source_key not in values or values[item.source_key] is None)
        )
        if missing:
            raise WorkbookWriteContractError(
                "required workbook mappings are missing: " + ", ".join(missing)
            )

        cached = load_workbook(source, data_only=True, keep_links=True)
        try:
            _verify_fixed_source_bindings(cached, profile, values)
        finally:
            cached.close()

        output.parent.mkdir(parents=True, exist_ok=True)
        if output.exists():
            raise WorkbookWriteContractError("output_path must not already exist")

        temporary: Path | None = None
        published = False
        try:
            workbook = load_workbook(source, data_only=False, keep_links=True)
            try:
                before_observation = _observe(workbook, profile, source.name)
                match_template_profile(
                    profile.template_profile, before_observation
                ).require_supported()
                self._verify_consumers(workbook, profile)
                before_cells = _cell_value_map(workbook)
                source_formula_cells = _formula_cells(workbook)

                applied_transformations: list[str] = []
                for binding in profile.write_bindings:
                    if binding.cell in source_formula_cells:
                        raise WorkbookWriteContractError(
                            f"source formula cell cannot be written: {binding.cell}"
                        )
                    sheet_name, coordinate = _split_cell(binding.cell)
                    if sheet_name not in workbook.sheetnames:
                        raise WorkbookStructuralRegressionError(
                            f"mapped worksheet missing: {sheet_name}"
                        )
                    workbook[sheet_name][coordinate] = _coerce_value(
                        binding, values.get(binding.source_key)
                    )

                for binding in profile.compatibility_bindings:
                    sheet_name, coordinate = _split_cell(binding.cell)
                    if sheet_name not in workbook.sheetnames:
                        raise WorkbookStructuralRegressionError(
                            f"compatibility worksheet missing: {sheet_name}"
                        )
                    surrogate = WorkbookWriteBinding(
                        cell=binding.cell,
                        source_key=binding.source_key,
                        value_kind=binding.value_kind,
                        required=binding.required,
                    )
                    workbook[sheet_name][coordinate] = _coerce_value(
                        surrogate, values.get(binding.source_key)
                    )
                    applied_transformations.append(binding.transformation_id)

                self._verify_consumers(workbook, profile)
                temporary = _new_owned_temporary(output)
                workbook.save(temporary)
            finally:
                workbook.close()

            _normalize_xlsx_zip(temporary)
            generated = load_workbook(temporary, data_only=False, keep_links=True)
            try:
                after_observation = _observe(generated, profile, output.name)
                match_template_profile(
                    profile.template_profile, after_observation
                ).require_supported()
                self._verify_consumers(generated, profile)
                after_cells = _cell_value_map(generated)
                after_formula_cells = _formula_cells(generated)
            finally:
                generated.close()

            expected_formula_cells = source_formula_cells - set(
                item.cell for item in profile.compatibility_bindings
            )
            missing_formulas = sorted(expected_formula_cells - after_formula_cells)
            if missing_formulas:
                raise WorkbookStructuralRegressionError(
                    "source formulas disappeared outside declared compatibility transformations: "
                    + ", ".join(missing_formulas)
                )

            changed = tuple(
                sorted(
                    key
                    for key in set(before_cells) | set(after_cells)
                    if before_cells.get(key) != after_cells.get(key)
                )
            )
            unexpected = sorted(set(changed) - set(profile.allowed_write_cells))
            if unexpected:
                raise WorkbookStructuralRegressionError(
                    "unexpected workbook cell changes outside profile allowlist: "
                    + ", ".join(unexpected)
                )

            if _sha256_file(source) != source_sha:
                raise WorkbookStructuralRegressionError(
                    "source workbook bytes changed during generation"
                )

            output_sha = _sha256_file(temporary)
            artifact = WorkbookGenerationArtifact(
                profile_id=profile.profile_id,
                profile_version=profile.profile_version,
                template_path=str(source),
                output_path=str(output),
                source_sha256=source_sha,
                output_sha256=output_sha,
                source_binding=source_binding,
                generated_at=generated_at,
                changed_cells=changed,
                applied_transformations=tuple(sorted(set(applied_transformations))),
                workbook_generated=True,
                excel_qualification_status="NOT_RUN",
            )

            _publish_new_artifact(temporary, output)
            published = True
            try:
                temporary.unlink()
            except OSError as exc:
                _cleanup_owned_after_publish_failure(temporary, output)
                raise WorkbookWriteContractError(
                    "owned temporary workbook could not be cleaned after publication"
                ) from exc
            temporary = None
            return artifact
        except Exception:
            if temporary is not None:
                if published:
                    _cleanup_owned_after_publish_failure(temporary, output)
                else:
                    try:
                        if temporary.exists():
                            temporary.unlink()
                    except OSError:
                        pass
            raise
