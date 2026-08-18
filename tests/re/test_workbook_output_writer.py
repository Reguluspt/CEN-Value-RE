from dataclasses import replace
from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.re.adapters.excel.fingerprint import UnsupportedTemplateError
from src.re.adapters.excel_output.n08_0038 import N08_0038_OUTPUT_PROFILE
from src.re.adapters.excel_output.openpyxl_writer import (
    OpenPyxlWorkbookOutputWriter,
    WorkbookSourceHashMismatchError,
    WorkbookWriteContractError,
)
from src.re.adapters.excel_output.profile import WorkbookValueKind
from src.re.ports.workbook_output import WorkbookGenerationSourceBinding


def _file_sha(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


def _synthetic_n08(path: Path, *, bad_g181=False, unknown_external=False):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for requirement in N08_0038_OUTPUT_PROFILE.template_profile.required_sheets:
        worksheet = workbook.create_sheet(requirement.name)
        worksheet.sheet_state = requirement.state.value
    for signature in N08_0038_OUTPUT_PROFILE.template_profile.formula_signatures:
        sheet, coordinate = signature.cell.rsplit("!", 1)
        workbook[sheet][coordinate] = signature.formula
    if bad_g181:
        workbook["Bangtinh"]["G181"] = "=1"
    workbook["Phieu TTTT"]["E5"] = "='[1]Nhập liệu'!F9"
    if unknown_external:
        workbook["Sheet1"]["Z99"] = "='[other.xlsx]Data'!A1"
    else:
        workbook["Sheet1"]["Z99"] = "UNKNOWN-SENTINEL"
    workbook.save(path)
    workbook.close()


def _values():
    values = {}
    for binding in (
        *N08_0038_OUTPUT_PROFILE.write_bindings,
        *N08_0038_OUTPUT_PROFILE.compatibility_bindings,
    ):
        if binding.value_kind is WorkbookValueKind.TEXT:
            values[binding.source_key] = "Tp. HCM" if binding.source_key == "subject.province" else "TEXT"
        elif binding.value_kind is WorkbookValueKind.FRACTION:
            values[binding.source_key] = "0.05"
        else:
            values[binding.source_key] = "123.45"
    return values


def _writer_for_source(source: Path):
    profile = replace(
        N08_0038_OUTPUT_PROFILE,
        source_exemplar_sha256=_file_sha(source),
    )
    return OpenPyxlWorkbookOutputWriter((profile,)), profile


def _binding():
    return WorkbookGenerationSourceBinding(
        case_id="case-1",
        final_valuation_snapshot_id="final-1",
        final_valuation_semantic_sha256="a" * 64,
    )


def test_writer_is_copy_on_write_allowlisted_and_deterministic(tmp_path):
    source = tmp_path / "source.xlsx"
    output_a = tmp_path / "generated-a.xlsx"
    output_b = tmp_path / "generated-b.xlsx"
    _synthetic_n08(source)
    source_before = source.read_bytes()
    writer, profile = _writer_for_source(source)

    artifact_a = writer.generate(
        profile_id=profile.profile_id,
        profile_version=profile.profile_version,
        template_path=str(source),
        output_path=str(output_a),
        values=_values(),
        source_binding=_binding(),
        generated_at="2026-08-18T08:00:00Z",
    )
    artifact_b = writer.generate(
        profile_id=profile.profile_id,
        profile_version=profile.profile_version,
        template_path=str(source),
        output_path=str(output_b),
        values=_values(),
        source_binding=_binding(),
        generated_at="2026-08-18T08:00:00Z",
    )

    assert source.read_bytes() == source_before
    assert artifact_a.output_sha256 == _file_sha(output_a)
    assert artifact_a.output_sha256 == artifact_b.output_sha256
    assert artifact_a.workbook_generated is True
    assert artifact_a.excel_qualification_status == "NOT_RUN"
    assert artifact_a.applied_transformations == ("localize-stale-phieu-tttt-e5",)
    assert set(artifact_a.changed_cells).issubset(profile.allowed_write_cells)
    assert "Phieu TTTT!E5" in artifact_a.changed_cells

    generated = load_workbook(output_a, data_only=False, keep_links=True)
    try:
        assert generated["Sheet1"]["Z99"].value == "UNKNOWN-SENTINEL"
        assert generated["Phieu TTTT"]["E5"].value == "Tp. HCM"
        for signature in profile.template_profile.formula_signatures:
            sheet, coordinate = signature.cell.rsplit("!", 1)
            assert generated[sheet][coordinate].value == signature.formula
        assert generated["Offical"]["E32"].value == "=Bangtinh!G181"
        assert generated["Bangtinh"]["G182"].value == "=ROUND(G181,-6)"
    finally:
        generated.close()


def test_writer_rejects_in_place_and_existing_output(tmp_path):
    source = tmp_path / "source.xlsx"
    _synthetic_n08(source)
    writer, profile = _writer_for_source(source)
    with pytest.raises(WorkbookWriteContractError, match="in place"):
        writer.generate(
            profile_id=profile.profile_id,
            profile_version=profile.profile_version,
            template_path=str(source),
            output_path=str(source),
            values=_values(),
            source_binding=_binding(),
            generated_at="2026-08-18T08:00:00Z",
        )

    output = tmp_path / "exists.xlsx"
    output.write_bytes(b"already-there")
    with pytest.raises(WorkbookWriteContractError, match="must not already exist"):
        writer.generate(
            profile_id=profile.profile_id,
            profile_version=profile.profile_version,
            template_path=str(source),
            output_path=str(output),
            values=_values(),
            source_binding=_binding(),
            generated_at="2026-08-18T08:00:00Z",
        )


def test_source_sha_and_structural_fingerprint_both_fail_closed(tmp_path):
    source = tmp_path / "source.xlsx"
    _synthetic_n08(source)
    writer = OpenPyxlWorkbookOutputWriter((N08_0038_OUTPUT_PROFILE,))
    with pytest.raises(WorkbookSourceHashMismatchError):
        writer.generate(
            profile_id=N08_0038_OUTPUT_PROFILE.profile_id,
            profile_version=N08_0038_OUTPUT_PROFILE.profile_version,
            template_path=str(source),
            output_path=str(tmp_path / "hash-fail.xlsx"),
            values=_values(),
            source_binding=_binding(),
            generated_at="2026-08-18T08:00:00Z",
        )

    mutated = tmp_path / "mutated.xlsx"
    _synthetic_n08(mutated, bad_g181=True)
    mutated_writer, mutated_profile = _writer_for_source(mutated)
    with pytest.raises(UnsupportedTemplateError):
        mutated_writer.generate(
            profile_id=mutated_profile.profile_id,
            profile_version=mutated_profile.profile_version,
            template_path=str(mutated),
            output_path=str(tmp_path / "fingerprint-fail.xlsx"),
            values=_values(),
            source_binding=_binding(),
            generated_at="2026-08-18T08:00:00Z",
        )


def test_unknown_external_dependency_fails_closed(tmp_path):
    source = tmp_path / "source.xlsx"
    _synthetic_n08(source, unknown_external=True)
    writer, profile = _writer_for_source(source)
    with pytest.raises(UnsupportedTemplateError):
        writer.generate(
            profile_id=profile.profile_id,
            profile_version=profile.profile_version,
            template_path=str(source),
            output_path=str(tmp_path / "external-fail.xlsx"),
            values=_values(),
            source_binding=_binding(),
            generated_at="2026-08-18T08:00:00Z",
        )
